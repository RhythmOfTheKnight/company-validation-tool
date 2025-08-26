import datetime
import logging 
import sys
import argparse
import pandas as pd
from tqdm import tqdm

from utils.file_utils import load_excel_file
from validators.batch_validator import validate_companies_batch
from api.postscodes import get_admin_district
from openpyxl.styles import PatternFill
from config import Columns, COLUMN_MAP
from datetime import datetime

def setup_logging():
    """Configure logging"""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s: %(message)s",
        datefmt = "%H:%M:%S",
        handlers = [
            logging.FileHandler("logs/datavalidation.log", mode = "a"),
            logging.StreamHandler(sys.stdout)
        ]
    )

    # Log date at the start
    logger = logging.getLogger(__name__)
    logger.info(f"SESSION STARTED: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    return logger

def update_company_data_from_api_results(df, logger):
    """
    Use the raw API data stored in API_RESULTS column to update
    original dataframe values. Where main dataupdating happens. 
    """
    logger.info("Enriching data from API results...")

    for idx, row in df.iterrows():
        api_data = row[Columns.API_DATA]
        if pd.notna(api_data) and isinstance(api_data, dict):
            for api_key, df_column_name in COLUMN_MAP.items():
                if api_key in api_data:
                    df.at[idx, df_column_name] = api_data[api_key]
    return df

def update_postcode_data_from_api_results(df, logger):        
    """
    Adds admin district using postcode
    """      
    logger.info("Updating with postcode data...")  

    if Columns.HEADQUARTERS not in df.columns:
         df[Columns.HEADQUARTERS] = None

    if Columns.PREVIOUS_HEADQUARTER_LOCATIONS not in df.columns:
        df[Columns.PREVIOUS_HEADQUARTER_LOCATIONS] = None

    for idx, row in tqdm(df.iterrows(), total=len(df)):
        postcode = row[Columns.POSTCODE]
        if pd.notna(postcode) and str(postcode).strip():
            new_district = get_admin_district(postcode)

            if new_district:
                current_district = row[Columns.HEADQUARTERS]

                if pd.notna(current_district) and current_district != new_district:

                    history = row[Columns.PREVIOUS_HEADQUARTER_LOCATIONS]
                    if pd.isna(history):
                        history = []

                    if not history or history[-1] != current_district:
                        history.append(current_district)
                        df.at[idx, Columns.PREVIOUS_HEADQUARTER_LOCATIONS] = history
                    

                df.at[idx, Columns.HEADQUARTERS] = new_district
    return df
    

def save_results_to_excel(df, output_file, sheet_name, logger):
    logger.info(f"Saving results to {output_file}...")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    with pd.ExcelWriter(output_file) as writer:
        df.to_excel(writer, sheet_name = sheet_name, index = False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        manual_review_col_idx = df.columns.get_loc(Columns.NEEDS_MANUAL_REVIEW)

        for row_idx, needs_review in enumerate(df[Columns.NEEDS_MANUAL_REVIEW], start=2): # start=2 for 1-based index and header
            if needs_review:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
    
    logger.info("Successfully saved and styled the output file.")


# Main Function

def main():
    parser = argparse.ArgumentParser(description="Validate companies via Companies House API")
    parser.add_argument("--input", required=True, help="Input Excel File: ")
    parser.add_argument("--sheet", required = True, help="Input Sheet Name: ")
    parser.add_argument("--output", required=True, help="Output Excel File: ")
    parser.add_argument("--api_key", required=True, help="Companies House API key: ")
    parser.add_argument("--limit", type=int, default=None, help="Max rows: ")
    args = parser.parse_args()

    logger = setup_logging()
    logger.info("Starting company validation…")

    # 1. Load Data
    input_df = load_excel_file(args.input, args.sheet)
    if input_df is None:
        logger.error(f"Failed to load data from {args.input}")
        return
    logger.info(f"Successfully loaded {len(input_df)} rows from {args.input}")
    if args.limit:
        input_df = input_df.head(args.limit)
        logger.info(f"Limited processing to {len(input_df)} rows")

    # 2. Validate the Data via companies house
    df_validated = validate_companies_batch(input_df, args.api_key)

    #3. Update the dataframe using companies and postcode helper functions
    df_updated = update_company_data_from_api_results(df_validated, logger)
    df_updated = update_postcode_data_from_api_results(df_updated, logger)

    # 4. Save the results
    if Columns.API_DATA in df_updated.columns:
        df_updated = df_updated.drop(columns=[Columns.API_DATA])
    
    save_results_to_excel(df_updated, args.output, args.sheet, logger)
    logger.info("Company validation process completed successfully.")


if __name__ == "__main__":
    main()
