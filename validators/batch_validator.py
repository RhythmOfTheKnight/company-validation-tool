import pandas as pd
import logging 
from datetime import datetime
from openpyxl.styles import PatternFill 
from openpyxl.utils import get_column_letter
from matchers.company_matcher import find_best_company_match
from openpyxl import load_workbook
import shutil
from config import COLUMN_MAP, ExcelStyles, Columns
from utils.company_type_utils import extract_company_type

# Set up Logger
logger = logging.getLogger(__name__)

RED_FILL = PatternFill(start_color=ExcelStyles.RED_FILL_COLOR,
                      end_color=ExcelStyles.RED_FILL_COLOR, 
                      fill_type="solid")


def load_excel_file(file_path, sheet_name = "main"):
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    logger.info(f"Loaded {len(df)} rows from {file_path} (sheet: {sheet_name})")
    return df

from datetime import datetime

def ensure_columns_exist(df):
    """Ensure all required output columns exist in the DataFrame."""
    # Create columns from COLUMN_MAP
    for col in COLUMN_MAP.values():
        if col not in df.columns:
            df[col] = ""
    
    # Create match metadata columns
    if Columns.MATCH_TYPE not in df.columns:
        df[Columns.MATCH_TYPE] = ""
    if Columns.CONFIDENCE not in df.columns:
        df[Columns.CONFIDENCE] = 0

def process_row(df, idx, api_key):
    """
    Enrich row `idx` of DataFrame `df` in-place, writing back all
    Companies House fields (including locality & postcode) via COLUMN_MAP,
    plus match metadata.
    Returns a small tracking dict.
    """
    # 1) Call the matcher
    row = df.loc[idx]
    company_data, mtype, conf, needs_manual = find_best_company_match(row, api_key)    # 2) Ensure output columns exist on df
    for col in COLUMN_MAP.values():
        if col not in df.columns:
            df[col] = ""
    if Columns.MATCH_TYPE not in df.columns:
        df[Columns.MATCH_TYPE] = ""
    if Columns.CONFIDENCE not in df.columns:
        df[Columns.CONFIDENCE] = 0

    # 3) Write back each standardized field
    if company_data:
        for std_key, col_name in COLUMN_MAP.items():
            df.at[idx, col_name] = company_data.get(std_key, "")
    else:
        # No company data found, but we can still extract company type from original data
        crn_value = row.get(Columns.CRN, "")
        company_name = row.get(Columns.CH_NAME, "") or row.get("Company Name", "")
        company_type = extract_company_type(crn_value=crn_value, company_name=company_name)
        df.at[idx, Columns.TYPE] = company_type

    # 4) Ensure Company Type is set even for matches (in case it wasn't extracted properly)
    if pd.isna(df.at[idx, Columns.TYPE]) or df.at[idx, Columns.TYPE] == "":
        crn_value = row.get(Columns.CRN, "")
        company_name = row.get(Columns.CH_NAME, "") or row.get("Company Name", "")
        company_type = extract_company_type(crn_value=crn_value, company_name=company_name)
        df.at[idx, Columns.TYPE] = company_type

    # 4) Overwrite Headquarters location if you want (optional)
    #    For example, if you want to copy registered locality into HQ:
    # df.at[idx, "Headquarters location"] = company_data.get("locality", row.get("Headquarters location",""))

    # 5) Write match metadata
    df.at[idx, Columns.MATCH_TYPE] = mtype
    df.at[idx, Columns.CONFIDENCE] = conf    # 6) Return tracking info
    return {
        "row_index": idx,
        "match_type": mtype,
        "confidence_score": conf,
        "needs_manual_review": needs_manual,
        "timestamp": datetime.now().isoformat()
    }

def process_all_rows(df, api_key, limit=None):
    """Process all rows and enrich with Companies House data."""
    
    # Ensure columns exist ONCE at the beginning
    ensure_columns_exist(df)
    
    tracking = []
    indices = list(df.index[:limit]) if limit else list(df.index)

    for idx in indices:
        logger.info(f"Processing row {idx+1}")
        rec = process_row(df, idx, api_key)
        tracking.append(rec)
    
    return tracking


def write_output(df, tracking, source_file, sheet_name, output_file):
    """
    1) Copy source_file to output_file (leaving source untouched)
    2) Open output_file in append mode, replacing only sheet_name
    3) Style rows needing manual review
    4) Append a report sheet
    """
    # 1) Copy the file
    shutil.copy(source_file, output_file)

    # 2) Write updated main sheet
    with pd.ExcelWriter(output_file,
                        engine="openpyxl",
                        mode="a",
                        if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 3) Style manual‐review rows
    wb = load_workbook(output_file)
    ws = wb[sheet_name]
    for item in tracking:
        if item["needs_manual_review"]:
            row = item["row_index"] + 2
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = RED_FILL
    wb.save(output_file)

    # 4) Append report sheet
    report_df = pd.DataFrame(tracking)
    with pd.ExcelWriter(output_file,
                        engine="openpyxl",
                        mode="a",
                        if_sheet_exists="replace") as writer:
        report_df.to_excel(writer, sheet_name="report", index=False)


def validate_companies_batch(source_file, api_key, sheet_name, output_file, limit=None):
    """
    Validate companies in a batch against Companies House API.

    Args:
        source_file (str): Path to the input Excel file (master.xlsx)
        api_key (str): Companies House API key
        sheet_name (str): Sheet name in the input Excel file (default: main)
        output_file (str): Path to the output Excel file (results.xlsx)
        limit (int, optional): Limit number of companies to process. Defaults to None (process all).

    Returns:
        pd.DataFrame: DataFrame with validated company data
        pd.DataFrame: DataFrame with tracking information
    """
    # Load the input file
    df = load_excel_file(source_file, sheet_name)

    # Apply limit to dataframe if specified
    if limit:
        df_limited = df.head(limit).copy()
    else:
        df_limited = df.copy()

    # Process rows and get tracking info
    tracking = process_all_rows(df_limited, api_key, limit)    # Write output to new file
    write_output(df_limited, tracking, source_file, sheet_name, output_file)

    return df_limited, pd.DataFrame(tracking)



